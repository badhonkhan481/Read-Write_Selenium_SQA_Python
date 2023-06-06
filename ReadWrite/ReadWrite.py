from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

print("sample test case started")
driver=webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.google.com/")

# driver.implicitly_wait(50)

src = "C:\\Users\\badhonkhan481\\Desktop\\SQA Project\\Read_Write_Selenium_Python\\Excel.xlsx"
workbook = openpyxl.load_workbook(src)
sheet = workbook.active
sheet = workbook['Sunday']

rowcount = sheet.max_row
colcount = sheet.max_column
print("rowcount:" + str(rowcount) + "colconut:" + str(colcount))

for i in range(2, rowcount+1):
    row=i+1
    key=sheet.cell(row, column=3).value

    p=driver.find_element("name", "q")
    p.send_keys(key)
    w = WebDriverWait(driver, 6)
    w.until(EC.presence_of_element_located((By.XPATH, "//ul")))
    # p.clear()
    time.sleep(2)
    ele = driver.find_element(By.XPATH,"//*[@id=\"Zrbbw\"]")
    ele2 = driver.find_element(By.XPATH, "//*[@id=\"vTtioc\"]")
    eleText = ele.text
    eleText2 = ele2.text
    print("Text value:" + eleText)
    sheet.cell(row, column=4).value=eleText
    sheet.cell(row, column=5).value = eleText2
    workbook.save(src)
    print("END OF WRITING DATA IN EXCEL")
    p.clear()

print("END OF WRITING DATA IN EXCEL")

def getText(ele):
    eleText = ele.text
    eleText = ""
    if eleText == "":
        eleText = ele.get_attribute("innerText")
        eleText = ""
        if eleText == "":
            eleText = ele.get_attribute("textContent")
    return eleText

def getText2(ele2):
    eleText2 = ele2.text
    eleText2 = ""
    if eleText2 == "":
        eleText2 = ele2.get_attribute("innerText")
        eleText2 = ""
        if eleText2 == "":
            eleText2 = ele2.get_attribute("textContent")
    return eleText2